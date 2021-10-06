import csv
import os
import os.path

import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

os.system("cls")





def generate_marksheet(grades_file,names_roll_file,subjects_master_file):
    DIRECTORY = "output"
    if not os.path.exists(DIRECTORY):
        os.makedirs(DIRECTORY)
    else:
        return
   
    # mapping a grade code to numerical value
    grade_map={"AA":10,"AB":9,"BB":8,"BC":7,"CC":6,"CD":5,"DD":4,"F":0,"I":0,"AA*":10,"AB*":9,"BB*":8,"BC*":7,"CC*":6,"CD*":5,"DD*":4,"F*":0,"I*":0}
    
    # mapping a course code to its [course name ,  L-T-P , credit]
    course_map={}
    with open(subjects_master_file,'r') as file:
        reader=csv.reader(file)
        counter=0
        for row in reader:
            counter+=1
            if(counter<=1):
                continue
            course_map[row[0]]=[row[1],row[2],row[3]]


    #finding the max semester number of a particular rolln number
    max_sem={}
    #add all the courses to a particular roll
    #1901ee12=[[sem_i,subcode_i,credit_i,grade_i,sub_type_i],[sem_j,subcode_j,credit_j,grade_j,sub_type_j]]
    roll_map={}
    with open(grades_file,'r') as file:
        reader=csv.reader(file)
        header=[]
        counter=0
        for row in reader:
            counter+=1
            if counter<=1:
                continue
            if row[0] in roll_map:
                roll_map[row[0]].append([row[1],row[2],int(row[3]),row[4].strip(),row[5]])
                max_sem[row[0]]=max(max_sem[row[0]],int(row[1]))
            else:
                roll_map[row[0]]=[]
                roll_map[row[0]].append([row[1],row[2],int(row[3]),row[4].strip(),row[5]])
                max_sem[row[0]]=int(row[1])
    
    # making file for each roll number and adding all sheets to them
    with open(names_roll_file,'r') as file:
        reader=csv.reader(file)
        counter=0
        for row in reader:
            counter+=1
            if(counter<=1):
                continue
            roll_path = os.path.join(DIRECTORY, row[0]+".xlsx")   
            roll_student=row[0]
            name_student=row[1]
            #appending the new Row to its course file path
            if not os.path.isfile(roll_path):
                wb = Workbook()
                sheet = wb.active
    
                sheet.append(["Roll No.",roll_student,"","","","","","",""])
                sheet.append(["Name of Student",name_student,"","","","","","",""])
                sheet.append(["Discipline",roll_student[4:6],"","","","","","",""])
                sem=["Semester No."]
                for x in range(max_sem[roll_student]):
                    sem.append(x+1)
                sheet.append(sem)
                sheet.title="Overall"
                wb.save(roll_path)
                for x in range(max_sem[roll_student]):
                    sheetname="Sem"+str(x+1)
                    wb.create_sheet(index=x+1,title=sheetname)
                    sheets=wb.sheetnames
                    sem_sheet=wb[sheets[x+1]]
                    sem_sheet.append(["Sl No.","Subject No.","Subject Name","L-T-P","Credit","Subject Type","Grade"])
                    wb.save(roll_path)


    #now adding details of courses for each semester in each roll_number xlxs file
    #also for each student store total credits taken and their corresponding results
    
    for roll in roll_map:
        roll_path = os.path.join(DIRECTORY, str(roll)+".xlsx")   
        wb=load_workbook(roll_path)
        sem_wise_credit_taken=["Semester wise Credit Taken"]
        SPI=["SPI"]
        for i in range(max_sem[roll]):
            sem_wise_credit_taken.append(0)
            SPI.append(0.0)
        for row in roll_map[roll]:
            semester_no=row[0]
            credits=row[2]
            grade=row[3]
            sheet=wb["Sem"+str(row[0])]
            row_count = sheet.max_row
            newRow=[row_count,row[1],course_map[row[1]][0],course_map[row[1]][1],credits,row[4],grade]
            sheet.append(newRow)
            sem_wise_credit_taken[int(semester_no)]+=int(credits)
            SPI[int(semester_no)]+=float(credits)*grade_map[grade]
            wb.save(roll_path)
        sheet=wb["Overall"]    
        total_credits_taken=["Total Credits Taken"]
        CPI=["CPI"]
        for i in range(max_sem[roll]):
            if(sem_wise_credit_taken[i+1]==0):
                SPI[i+1]=0.00
            else:
                SPI[i+1]/=sem_wise_credit_taken[i+1]*1.0
            if(i==0):
                total_credits_taken.append(sem_wise_credit_taken[1])
                CPI.append(SPI[i+1])
            else:
                total_credits_taken.append(sem_wise_credit_taken[i+1]+total_credits_taken[i])
                if(total_credits_taken[i+1]==0):
                    CPI.append(0.0)
                else:
                    CPI.append((CPI[i]*total_credits_taken[i] +SPI[i+1]*sem_wise_credit_taken[i+1])/total_credits_taken[i+1])
        sheet.append(sem_wise_credit_taken)
        for i in range(max_sem[roll]):
            SPI[i+1]=round(SPI[i+1],2)
        sheet.append(SPI)
        sheet.append(total_credits_taken)
        for i in range(max_sem[roll]):
            CPI[i+1]=round(CPI[i+1],2)
        sheet.append(CPI)
        wb.save(roll_path)
    return








#storing file addresses for each csv file
grades_file="grades.csv"
names_roll_file="names-roll.csv"
subjects_master_file="subjects_master.csv"


#task
generate_marksheet(grades_file,names_roll_file,subjects_master_file)
