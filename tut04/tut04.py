import csv
import os
import os.path
from openpyxl import Workbook
from openpyxl import load_workbook


os.system("cls")

def output_by_subject(csv_file):
    DIRECTORY = "output_by_subject"

    if not os.path.exists(DIRECTORY):
        os.makedirs(DIRECTORY)
    
    with open(csv_file,'r') as file:
        reader=csv.reader(file)
    
        counter=0
        header=[]
        for row in reader:
            counter+=1
            if counter==1:
                header=row
            else:
                course_path = os.path.join(DIRECTORY, row[3]+".xlsx")   
                #appending the new Row to its course file path
                if os.path.isfile(course_path):
                    wb = load_workbook(course_path)
                    sheet = wb.active
                    sheet.append([row[0],row[1],row[3],row[8]])
                    wb.save(course_path)     
                #creating a file with the course code and writing header row and new row
                else:
                    wb = Workbook()
                    sheet = wb.active
                    sheet.append(['rollno','register_sem','subno','sub_type'])
                    sheet.append([row[0],row[1],row[3],row[8]])     
                    wb.save(course_path)    
    return

def output_individual_roll(csv_file):
    DIRECTORY = "output_individual_roll"

    if not os.path.exists(DIRECTORY):
        os.makedirs(DIRECTORY)
    
    with open(csv_file,'r') as file:
        reader=csv.reader(file)
    
        counter=0
        header=[]
        for row in reader:
            counter+=1
            if counter==1:
                header=row
            else:
                roll_path = os.path.join(DIRECTORY, row[0]+".xlsx")   
                #appending the new Row to its roll path
                if os.path.isfile(roll_path):
                    wb = load_workbook(roll_path)
                    sheet = wb.active
                    sheet.append([row[0],row[1],row[3],row[8]])
                    wb.save(roll_path)     
                #creating a file with the roll code and writing header row and new row
                else:
                    wb = Workbook()
                    sheet = wb.active
                    sheet.append(['rollno','register_sem','subno','sub_type'])
                    sheet.append([row[0],row[1],row[3],row[8]])     
                    wb.save(roll_path)
    return


csv_file="regtable_old.csv"

#task1
output_individual_roll(csv_file)

#task2
output_by_subject(csv_file)

